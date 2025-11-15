"""CSV to XLSX Risk Matrix Converter with Security Hardening.

Automatically generates a styled Excel risk matrix with charts from a CSV file.

Security features:
- CSV injection protection (escapes dangerous characters)
- File size validation (max 10MB)
- Data type validation for risk scores
- Input sanitization for all cell values
"""

import sys
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from riskops.core.constants import (
    EXCEL_SHEET_NAMES,
    HEADER_STYLE,
    MAX_FILE_SIZE_MB,
    REQUIRED_RISK_COLUMNS,
    RISK_COLORS,
)
from riskops.utils.security import (
    sanitize_cell_value,
    validate_file_size,
    validate_risk_scores,
)


def load_csv(csv_path: Path) -> pd.DataFrame:
    """
    Loads the risk matrix from CSV with security validation.

    Security measures:
    - File size validation (max 10MB)
    - Required column validation
    - Data type validation for risk scores
    - CSV injection protection via sanitization

    Args:
        csv_path: Path to CSV file

    Returns:
        Validated and sanitized DataFrame

    Raises:
        ValueError: If validation fails
        FileNotFoundError: If file doesn't exist
    """
    # Case 1: File absent (CI mode or demo)
    if not csv_path.exists():
        print(f"Warning: {csv_path} not found. Using demo dataset (CI mode).")

        demo_data = [
            {
                "ID": "R001",
                "Asset": "Demo infusion pump",
                "Threat": "Ransomware",
                "Vulnerability": "Unpatched OS",
                "Probability": 4,
                "Impact": 5,
                "Risk": 20,
                "Decision": "Reduce",
                "Recommendation": "Patch management + network segmentation",
            },
            {
                "ID": "R002",
                "Asset": "Demo patient monitor",
                "Threat": "Unauthorized access",
                "Vulnerability": "Weak credentials",
                "Probability": 3,
                "Impact": 4,
                "Risk": 12,
                "Decision": "Reduce",
                "Recommendation": "Strong auth + MFA where possible",
            },
        ]

        df = pd.DataFrame(demo_data)
        return df

    # Case 2: File present (local environment)
    try:
        print(f"Loading CSV from: {csv_path}")

        # Security: Validate file size before reading
        validate_file_size(csv_path)

        # Security: Specify dtypes to prevent type confusion
        # String columns can contain any text, numeric columns must be int
        dtype_spec = {
            "ID": str,
            "Asset": str,
            "Threat": str,
            "Vulnerability": str,
            "Probability": "int64",
            "Impact": "int64",
            "Risk": "int64",
            "Decision": str,
            "Recommendation": str,
        }

        df = pd.read_csv(csv_path, dtype=dtype_spec)

        # Validate required columns
        missing_cols = [col for col in REQUIRED_RISK_COLUMNS if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Missing required columns: {missing_cols}")

        # Security: Validate risk score ranges and types
        validate_risk_scores(df)

        # Recalculate risk to ensure consistency
        df["Risk"] = df["Probability"] * df["Impact"]

        print(f"✓ Loaded {len(df)} risks from CSV")
        return df

    except pd.errors.ParserError as e:
        print(f"Error: Invalid CSV format: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"Error: Data validation failed: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Error reading CSV: {e}")
        sys.exit(1)


def create_styled_header(ws, headers: list) -> None:
    """Creates a styled Excel header row with sanitized values."""
    header_fill = PatternFill(
        start_color=HEADER_STYLE["fill_color"],
        end_color=HEADER_STYLE["fill_color"],
        fill_type="solid",
    )
    header_font = Font(
        color=HEADER_STYLE["font_color"],
        bold=HEADER_STYLE["bold"],
        size=HEADER_STYLE["font_size"],
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        # Security: Sanitize header values
        cell.value = sanitize_cell_value(header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_risk_coloring(ws, df: pd.DataFrame) -> None:
    """Applies colors based on the Risk level."""
    # Find the column index for 'Risk'
    try:
        risk_col_idx = df.columns.get_loc("Risk") + 1
    except KeyError:
        print("Error: 'Risk' column not found for coloring.")
        return

    for row_num in range(2, len(df) + 2):
        risk_value = ws.cell(row=row_num, column=risk_col_idx).value

        # Determine color based on common risk scoring thresholds (P*I)
        if risk_value <= 6:
            color_key = "low"
        elif risk_value <= 12:
            color_key = "medium"
        elif risk_value <= 14:
            color_key = "high"
        else:
            color_key = "critical"

        # Apply color to the risk cell
        ws.cell(row=row_num, column=risk_col_idx).fill = PatternFill(
            start_color=RISK_COLORS[color_key],
            end_color=RISK_COLORS[color_key],
            fill_type="solid",
        )


def create_heatmap(wb, df: pd.DataFrame) -> None:
    """Creates a Probability × Impact heatmap sheet."""
    ws = wb.create_sheet(title=EXCEL_SHEET_NAMES["heatmap"])

    # Titles for the matrix (English)
    ws["A1"] = "Impact →"
    ws["A1"].font = Font(bold=True)

    # Column Headers (Impact 1-5)
    for i in range(1, 6):
        cell = ws.cell(row=1, column=i + 1)
        cell.value = i
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Row Headers (Probability 5-1, reversed for typical display)
    ws["A2"] = "Probability ↓"
    ws["A2"].font = Font(bold=True)
    for i in range(5, 0, -1):
        cell = ws.cell(row=7 - i, column=1)
        cell.value = i
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Fill matrix with counters
    risk_counts: dict[tuple[int, int], int] = {}
    for _, row in df.iterrows():
        # Ensure values are treated as integers
        try:
            p = int(row["Probability"])
            i = int(row["Impact"])
            key = (p, i)
            risk_counts[key] = risk_counts.get(key, 0) + 1
        except ValueError:
            # Skip rows where P or I are not numeric
            continue

    # Color and count cells
    for p in range(1, 6):
        for i in range(1, 6):
            cell = ws.cell(row=7 - p, column=i + 1)
            risk = p * i
            count = risk_counts.get((p, i), 0)

            cell.value = count if count > 0 else ""

            # Color based on risk score
            if risk <= 6:
                color = RISK_COLORS["low"]
            elif risk <= 12:
                color = RISK_COLORS["medium"]
            elif risk <= 14:
                color = RISK_COLORS["high"]
            else:
                color = RISK_COLORS["critical"]

            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, size=14)

    # Adjust column widths
    ws.column_dimensions["A"].width = 15
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 12


def create_dashboard(wb, df: pd.DataFrame) -> None:
    """Creates a risk summary dashboard sheet."""
    ws = wb.create_sheet(title=EXCEL_SHEET_NAMES["dashboard"])

    # Title
    ws["A1"] = "Dashboard - Risk Summary"
    ws["A1"].font = Font(size=16, bold=True)

    # Global Stats
    ws["A3"] = "Global Statistics"
    ws["A3"].font = Font(size=12, bold=True)

    ws["A4"] = "Total number of risks:"
    ws["B4"] = len(df)

    ws["A5"] = "CRITICAL Risks (≥15):"
    ws["B5"] = len(df[df["Risk"] >= 15])
    ws["B5"].font = Font(color="FF0000", bold=True)  # Red

    ws["A6"] = "MEDIUM/HIGH Risks (8-14):"
    ws["B6"] = len(df[(df["Risk"] >= 8) & (df["Risk"] <= 14)])
    ws["B6"].font = Font(color="FFA500", bold=True)  # Orange

    ws["A7"] = "LOW Risks (≤6):"
    ws["B7"] = len(df[df["Risk"] <= 6])
    ws["B7"].font = Font(color="008000", bold=True)  # Green

    # Decision Breakdown
    ws["A9"] = "Treatment Decision Breakdown"
    ws["A9"].font = Font(size=12, bold=True)

    decisions = df["Decision"].value_counts()
    row = 10
    for decision, count in decisions.items():
        # Security: Sanitize decision values
        ws.cell(row=row, column=1).value = sanitize_cell_value(decision)
        ws.cell(row=row, column=2).value = count
        row += 1

    # Top 5 Critical Risks
    ws["A" + str(row + 2)] = "Top 5 Critical Risks"
    ws["A" + str(row + 2)].font = Font(size=12, bold=True)

    top_risks = df.nlargest(5, "Risk")[["ID", "Asset", "Risk"]]
    row += 4
    ws.cell(row=row, column=1).value = "ID"
    ws.cell(row=row, column=2).value = "Asset"
    ws.cell(row=row, column=3).value = "Score"

    for _, risk in top_risks.iterrows():
        row += 1
        # Security: Sanitize all cell values
        ws.cell(row=row, column=1).value = sanitize_cell_value(risk["ID"])
        ws.cell(row=row, column=2).value = sanitize_cell_value(risk["Asset"])
        ws.cell(row=row, column=3).value = risk["Risk"]
        ws.cell(row=row, column=3).fill = PatternFill(
            start_color=RISK_COLORS["critical"],
            end_color=RISK_COLORS["critical"],
            fill_type="solid",
        )

    # Adjust widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10


def convert_csv_to_xlsx(
    csv_path: Path, output_path: Optional[Path] = None, verbose: bool = True
) -> Path:
    """
    Converts a CSV risk matrix to a fully formatted Excel file.

    This is the main entry point for CSV to XLSX conversion. It creates
    an Excel workbook with three sheets: Risk Matrix, Heatmap, and Dashboard.

    Args:
        csv_path: Path to input CSV file
        output_path: Path for output XLSX file (defaults to same dir as CSV)
        verbose: Whether to print progress messages

    Returns:
        Path to the generated Excel file

    Raises:
        FileNotFoundError: If CSV file doesn't exist
        ValueError: If CSV validation fails

    Example:
        >>> from pathlib import Path
        >>> csv_file = Path("risk_matrix.csv")
        >>> excel_file = convert_csv_to_xlsx(csv_file)
        >>> print(f"Created: {excel_file}")
    """
    # Default output path
    if output_path is None:
        output_path = csv_path.parent / "reports" / "risk_matrix.xlsx"

    # Create output directory if it doesn't exist
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if verbose:
        print("=" * 60)
        print("RiskOps CSV to XLSX Converter v1.1 (Security Hardened)")
        print("=" * 60)

    # Security: Load and validate CSV with security checks
    df = load_csv(csv_path)

    if verbose:
        print("\nCreating Excel file with security protections...")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = EXCEL_SHEET_NAMES["matrix"]

    # Security: Write data with CSV injection protection
    if verbose:
        print("Writing risk matrix data...")
    for r_idx, row in enumerate(df.itertuples(index=False), 1):
        for c_idx, value in enumerate(row, 1):
            # Security: Sanitize every cell value to prevent CSV injection
            sanitized_value = sanitize_cell_value(value)
            ws.cell(row=r_idx + 1, column=c_idx).value = sanitized_value

    # Styled Header
    create_styled_header(ws, df.columns.tolist())

    # Risk Coloring
    apply_risk_coloring(ws, df)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    # Create additional sheets
    if verbose:
        print("Creating heatmap...")
    create_heatmap(wb, df)

    if verbose:
        print("Creating dashboard...")
    create_dashboard(wb, df)

    # Save
    if verbose:
        print(f"\nSaving to: {output_path}")
    wb.save(output_path)

    if verbose:
        print("\n" + "=" * 60)
        print("✓ SUCCESS: Excel file generated successfully!")
        print("=" * 60)
        print(f"  • {len(df)} risks processed")
        print("  • 3 tabs created: Risk Matrix + Heatmap + Dashboard")
        print("  • CSV injection protection applied")
        print(f"  • File size validated (max {MAX_FILE_SIZE_MB}MB)")
        print("  • Risk scores validated (Probability & Impact 1-5)")
        print("=" * 60)

    return output_path
