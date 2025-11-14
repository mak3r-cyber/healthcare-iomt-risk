#!/usr/bin/env python3
"""CSV to XLSX Risk Matrix Converter
Automatically generates a styled Excel risk matrix with charts from a CSV file.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import sys
from pathlib import Path

csv_path = "02-Matrices/risk_matrix.csv"


# Colors for Risk Scoring (English Keys)
COLORS = {
    "low": "C6EFCE",  # Light Green
    "medium": "FFEB9C",  # Yellow
    "high": "FFC7CE",  # Light Red
    "critical": "FF0000",  # Red
}


def load_csv(csv_path: Path) -> pd.DataFrame:
    """
    Loads the risk matrix from CSV.

    - En local : lit 02-Matrices/risk_matrix.csv
    - En CI (GitHub Actions) : si le fichier est absent (gitignore), génère un dataset de démo
    """
    # Cas 1 : fichier absent (runner CI ou repo public sans CSV réel)
    if not csv_path.exists():
        print(f"Warning: {csv_path} not found. Using demo dataset (CI mode).")

        demo_data = [
            {
                "ID": "R001",
                "Asset": "Demo infusion pump",
                "Threat": "Ransomware",
                "Vulnerability": "Unpatched OS",
                "Probability": 4,
                "Impact": 5,
                "Risk": 20,
                "Decision": "Reduce",
                "Recommendation": "Patch management + network segmentation",
            },
            {
                "ID": "R002",
                "Asset": "Demo patient monitor",
                "Threat": "Unauthorized access",
                "Vulnerability": "Weak credentials",
                "Probability": 3,
                "Impact": 4,
                "Risk": 12,
                "Decision": "Reduce",
                "Recommendation": "Strong auth + MFA where possible",
            },
        ]

        df = pd.DataFrame(demo_data)
        return df

    # Cas 2 : fichier présent (ton environnement local)
    try:
        print(f"Loading CSV from: {csv_path}")
        df = pd.read_csv(csv_path)

        required_cols = [
            "ID",
            "Asset",
            "Threat",
            "Vulnerability",
            "Probability",
            "Impact",
            "Risk",
            "Decision",
            "Recommendation",
        ]

        if not all(col in df.columns for col in required_cols):
            print(f"Error: Missing columns in CSV. Expected: {required_cols}")
            sys.exit(1)

        # Recalcule le risque pour être sûr
        df["Risk"] = df["Probability"] * df["Impact"]

        return df
    except Exception as e:
        print(f"Error reading CSV: {e}")
        sys.exit(1)


def create_styled_header(ws, headers: list):
    """Creates a styled Excel header row."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_risk_coloring(ws, df: pd.DataFrame):
    """Applies colors based on the Risk level."""
    # Find the column index for 'Risk'
    try:
        risk_col_idx = df.columns.get_loc("Risk") + 1
    except KeyError:
        # Should not happen if load_csv succeeded, but for safety
        print("Error: 'Risk' column not found for coloring.")
        return

    for row_num in range(2, len(df) + 2):
        risk_value = ws.cell(row=row_num, column=risk_col_idx).value

        # Determine color based on common risk scoring thresholds (P*I)
        if risk_value <= 6:
            color_key = "low"
        elif risk_value <= 12:
            color_key = "medium"
        elif risk_value <= 14:
            color_key = "high"
        else:
            color_key = "critical"

        # Apply color to the risk cell
        ws.cell(row=row_num, column=risk_col_idx).fill = PatternFill(
            start_color=COLORS[color_key], end_color=COLORS[color_key], fill_type="solid"
        )


def create_heatmap(wb, df: pd.DataFrame):
    """Creates a Probability × Impact heatmap sheet."""
    ws = wb.create_sheet(title="Heatmap")

    # Titles for the matrix (English)
    ws["A1"] = "Impact →"
    ws["A1"].font = Font(bold=True)

    # Column Headers (Impact 1-5)
    for i in range(1, 6):
        cell = ws.cell(row=1, column=i + 1)
        cell.value = i
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Row Headers (Probability 5-1, reversed for typical display)
    ws["A2"] = "Probability ↓"
    ws["A2"].font = Font(bold=True)
    for i in range(5, 0, -1):
        cell = ws.cell(row=7 - i, column=1)
        cell.value = i
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Fill matrix with counters
    risk_counts = {}
    for _, row in df.iterrows():
        # Ensure values are treated as integers
        try:
            p = int(row["Probability"])
            i = int(row["Impact"])
            key = (p, i)
            risk_counts[key] = risk_counts.get(key, 0) + 1
        except ValueError:
            # Skip rows where P or I are not numeric
            continue

    # Color and count cells
    for p in range(1, 6):
        for i in range(1, 6):
            cell = ws.cell(row=7 - p, column=i + 1)
            risk = p * i
            count = risk_counts.get((p, i), 0)

            cell.value = count if count > 0 else ""

            # Color based on risk score
            if risk <= 6:
                color = COLORS["low"]
            elif risk <= 12:
                color = COLORS["medium"]
            elif risk <= 14:
                color = COLORS["high"]
            else:
                color = COLORS["critical"]

            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, size=14)

    # Adjust column widths
    ws.column_dimensions["A"].width = 15
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 12


def create_dashboard(wb, df: pd.DataFrame):
    """Creates a risk summary dashboard sheet."""
    ws = wb.create_sheet(title="Dashboard")

    # Title
    ws["A1"] = "Dashboard - Risk Summary"
    ws["A1"].font = Font(size=16, bold=True)

    # Global Stats
    ws["A3"] = "Global Statistics"
    ws["A3"].font = Font(size=12, bold=True)

    ws["A4"] = "Total number of risks:"
    ws["B4"] = len(df)

    ws["A5"] = "CRITICAL Risks (≥15):"
    ws["B5"] = len(df[df["Risk"] >= 15])
    ws["B5"].font = Font(color="FF0000", bold=True)  # Red

    ws["A6"] = "MEDIUM/HIGH Risks (8-14):"
    ws["B6"] = len(df[(df["Risk"] >= 8) & (df["Risk"] <= 14)])
    ws["B6"].font = Font(color="FFA500", bold=True)  # Orange

    ws["A7"] = "LOW Risks (≤6):"
    ws["B7"] = len(df[df["Risk"] <= 6])
    ws["B7"].font = Font(color="008000", bold=True)  # Green

    # Decision Breakdown
    ws["A9"] = "Treatment Decision Breakdown"
    ws["A9"].font = Font(size=12, bold=True)

    decisions = df["Decision"].value_counts()
    row = 10
    for decision, count in decisions.items():
        ws.cell(row=row, column=1).value = decision
        ws.cell(row=row, column=2).value = count
        row += 1

    # Top 5 Critical Risks
    ws["A" + str(row + 2)] = "Top 5 Critical Risks"
    ws["A" + str(row + 2)].font = Font(size=12, bold=True)

    top_risks = df.nlargest(5, "Risk")[["ID", "Asset", "Risk"]]
    row += 4
    ws.cell(row=row, column=1).value = "ID"
    ws.cell(row=row, column=2).value = "Asset"
    ws.cell(row=row, column=3).value = "Score"

    for _, risk in top_risks.iterrows():
        row += 1
        ws.cell(row=row, column=1).value = risk["ID"]
        ws.cell(row=row, column=2).value = risk["Asset"]
        ws.cell(row=row, column=3).value = risk["Risk"]
        ws.cell(row=row, column=3).fill = PatternFill(
            start_color=COLORS["critical"], end_color=COLORS["critical"], fill_type="solid"
        )

    # Adjust widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10


def main():
    """Main function to execute script."""
    # Paths (using correct project structure and English names)
    project_root = Path(__file__).parent.parent
    csv_path = project_root / "02-Matrices" / "risk_matrix.csv"
    xlsx_path = project_root / "docs" / "reports" / "risk_matrix.xlsx"

    # Create output directory if it doesn't exist
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Loading CSV: {csv_path}")
    df = load_csv(csv_path)

    print("Creating Excel file...")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk_Matrix"

    # Write data (starting from row 2, index=False is for Pandas' itertuples)
    for r_idx, row in enumerate(df.itertuples(index=False), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx).value = value

    # Styled Header
    create_styled_header(ws, df.columns.tolist())

    # Risk Coloring
    apply_risk_coloring(ws, df)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    # Create additional sheets
    create_heatmap(wb, df)
    create_dashboard(wb, df)

    # Save
    print(f"Saving to: {xlsx_path}")
    wb.save(xlsx_path)

    print("Creating Excel file...")  # Pas besoin de f-string ici
    print("Success: Excel file generated successfully!")  # Pas besoin de f-string ici
    print(
        " - " + str(len(df)) + " risks processed"
    )  # F-string nécessaire ici pour afficher len(df)
    print("  - 3 tabs created: Matrix + Heatmap + Dashboard")  # Pas besoin de f-string ici


if __name__ == "__main__":
    main()
